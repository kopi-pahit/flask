from flask import Flask, render_template, request, redirect, url_for, session, send_file, jsonify, abort
import sqlite3
import os
from openpyxl import Workbook
from datetime import datetime, timedelta
import json

app = Flask(__name__)
app.secret_key = "secret_key"
DATABASE = 'data.db'


DATA_DIR = os.path.join(os.path.dirname(__file__), 'quran')

@app.route('/search', methods=['GET', 'POST'])
def search():
    if request.method == 'POST':
        search_query = request.form['search_query']
        results_absen = search_in_database('absen', search_query)
        results_hafalan = search_in_database('hafalan', search_query)
        return render_template('search_results.html', results_absen=results_absen, results_hafalan=results_hafalan)
    return render_template('dashboard.html')

def search_in_database(table_name, query):
    with sqlite3.connect('data.db') as conn:
        c = conn.cursor()
        c.execute(f"SELECT * FROM {table_name} WHERE nama LIKE ?", ('%' + query + '%',))
        results = c.fetchall()
    return results


# ambil semua data absen yang memiliki tanggal yang sama dengan tanggal terbaru
def get_latest_absen_data():
    today = datetime.date.today()
    query = "SELECT * FROM absen WHERE tgl = ?;"
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.cursor()
        cursor.execute(query, (today,))
        data = cursor.fetchall()
    return data

# ambil semua data hafalan yang memiliki tanggal yang sama dengan tanggal terbaru
def get_latest_hafalan_data():
    today = datetime.date.today()
    query = "SELECT * FROM hafalan WHERE tgl = ?;"
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.cursor()
        cursor.execute(query, (today,))
        data = cursor.fetchall()
    return data

def get_absen_data(start_date_obj, end_date_obj):
    query = "SELECT * FROM absen WHERE tgl BETWEEN ? AND ?;"
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.cursor()
        cursor.execute(query, (start_date_obj, end_date_obj))
        data = cursor.fetchall()
    return data

def get_hafalan_data(start_date_obj=None, end_date_obj=None):
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.cursor()
        if start_date_obj and end_date_obj:
            query = "SELECT * FROM hafalan WHERE tgl BETWEEN ? AND ?;"
            cursor.execute(query, (start_date_obj, end_date_obj))
        else:
            query = "SELECT * FROM hafalan;"
            cursor.execute(query)
        data = cursor.fetchall()
    return data



def get_weekly_data(table_name):
    today = datetime.date.today()
    week_ago = today - datetime.timedelta(days=7)
    query = f"SELECT * FROM {table_name} WHERE tgl BETWEEN ? AND ?;"
    
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.cursor()
        cursor.execute(query, (week_ago, today))
        data = cursor.fetchall()
    
    return data

# Fungsi untuk menjalankan query SQLite
def run_query(query, parameters=()):
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.cursor()
        cursor.execute(query, parameters)
        conn.commit()
        return cursor

# Membuat tabel pengguna jika belum ada
def create_users_table():
    query = '''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        password TEXT
    );
    '''
    run_query(query)

# Menambahkan pengguna baru
def add_user(username, password):
    query = "INSERT INTO users (username, password) VALUES (?, ?);"
    run_query(query, (username, password))

# Memeriksa kredensial pengguna saat login
def check_credentials(username, password):
    query = "SELECT * FROM users WHERE username=? AND password=?;"
    result = run_query(query, (username, password)).fetchone()
    return result is not None

# Fungsi untuk mengecek apakah pengguna sudah login
def is_logged_in():
    return 'username' in session

## hari ##
DAYS_IN_INDO = {
    "Monday": "Senin",
    "Tuesday": "Selasa",
    "Wednesday": "Rabu",
    "Thursday": "Kamis",
    "Friday": "Jumat",
    "Saturday": "Sabtu",
    "Sunday": "Minggu"
}

@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    page = request.args.get('page', 1, type=int)  # Mendapatkan halaman saat ini, default ke 1
    per_page = 5  # Menampilkan 5 baris per halaman

    start_date = request.args.get('start_date', '')  # Mengambil tanggal mulai dari query string
    end_date = request.args.get('end_date', '')  # Mengambil tanggal akhir dari query string
    
    now = datetime.now()
    day_of_week_english = now.strftime("%A")  # Get the day of the week in English
    current_date_in_indo = DAYS_IN_INDO.get(day_of_week_english, day_of_week_english)
    formatted_date = now.strftime("%d %B %Y")  # Format the date (optional)

    if request.method == 'POST':
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        
        # Mengubah string tanggal menjadi objek datetime
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        
        # Menyimpan tanggal dalam session untuk digunakan di GET
        session['start_date'] = start_date
        session['end_date'] = end_date
        
         # Mengambil data absen berdasarkan rentang tanggal
        absen_data = get_absen_data(start_date_obj, end_date_obj)
        hafalan_data = get_hafalan_data(start_date_obj, end_date_obj)
        
        # Paginate data
        total_absen = len(absen_data)
        paginated_absen_data = absen_data[(page - 1) * per_page: page * per_page]
        
        total_hafalan = len(hafalan_data)
        paginated_hafalan_data = hafalan_data[(page - 1) * per_page: page * per_page]

        return render_template('dashboard.html', 
                               absen_data=paginated_absen_data, 
                               hafalan_data=paginated_hafalan_data,
                               total_absen=total_absen, 
                               total_hafalan=total_hafalan,
                               per_page=per_page, 
                               page=page,
                               start_date=start_date,
                               end_date=end_date, current_date=current_date_in_indo, formatted_date=formatted_date)
    else:
        start_date = session.get('start_date', '')
        end_date = session.get('end_date', '')
        
        if start_date and end_date:
            # Mengubah string tanggal menjadi objek datetime
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            
            # Mengambil data absen berdasarkan rentang tanggal
            absen_data = get_absen_data(start_date_obj, end_date_obj)
            hafalan_data = get_hafalan_data(start_date_obj, end_date_obj)  # Ambil data hafalan, tambahkan filter jika diperlukan
            
            # Paginate data
            total_absen = len(absen_data)
            paginated_absen_data = absen_data[(page - 1) * per_page: page * per_page]
            
            total_hafalan = len(hafalan_data)
            paginated_hafalan_data = hafalan_data[(page - 1) * per_page: page * per_page]
        else:
            absen_data = []
            hafalan_data = []
            total_absen = 0
            total_hafalan = 0
            paginated_absen_data = []
            paginated_hafalan_data = []

        return render_template('dashboard.html', 
                               absen_data=paginated_absen_data, 
                               hafalan_data=paginated_hafalan_data,
                               total_absen=total_absen, 
                               total_hafalan=total_hafalan,
                               per_page=per_page, 
                               page=page,
                               start_date=start_date,
                               end_date=end_date, current_date=current_date_in_indo, formatted_date=formatted_date)


    # Ambil data terbaru
    #absen_data = get_latest_absen_data()
    #hafalan_data = get_latest_hafalan_data()
    #return render_template('dashboard.html', absen_data=absen_data, hafalan_data=hafalan_data)

@app.route('/quran')
def quran():
    if not is_logged_in():
        return redirect(url_for('login'))

    surahs = []
    try:
        file_list = os.listdir(DATA_DIR)
        for filename in file_list:
            if filename.endswith('.json'):
                surah_number = int(filename.split('.')[0])
                file_path = os.path.join(DATA_DIR, filename)
                with open(file_path, 'r', encoding='utf-8') as f:
                    surah_data = json.load(f)
                    if str(surah_number) in surah_data:
                        surah_name = surah_data[str(surah_number)]['name']
                        surahs.append({'number': surah_number, 'name': surah_name})
    except Exception as e:
        print(f"Error reading surah files: {e}")

    surahs = sorted(surahs, key=lambda x: x['number'])
    
    return render_template('quran.html', surahs=surahs)

@app.route('/surah/<int:surah_number>')
def surah(surah_number):
    if not is_logged_in():
        return redirect(url_for('login'))

    surah_file = os.path.join(DATA_DIR, f'{surah_number}.json')
    if os.path.exists(surah_file):
        with open(surah_file, 'r', encoding='utf-8') as f:
            surah_data = json.load(f)
            surah_info = surah_data.get(str(surah_number), {})
            surah_name = surah_info.get('name', 'Unknown Surah')
            surah_ayahs = surah_info.get('text', {})
            translations = surah_info.get('translations', {}).get('id', {}).get('text', {})
            surah_tafsirs = surah_info.get('tafsir', {}).get('id', {})
            return render_template('baca.html', surah_number=surah_number, surah_name=surah_name, surah_ayahs=surah_ayahs, translations=translations, surah_tafsirs=surah_tafsirs)
    else:
        abort(404, description="Surah not found")

#########################################################################################


@app.route('/')
def index():
    if is_logged_in():
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))
    
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if check_credentials(username, password):
            session['username'] = username
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error='Username atau password salah')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        add_user(username, password)
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('index'))

@app.route('/absen')
def absen_form():
    return render_template('absensi.html')

# Fungsi untuk menambahkan data absen ke dalam database
@app.route('/add_absen', methods=['POST'])
def add_absen():
    if request.method == 'POST':
        tgl = request.form.get('tgl')
        nama = request.form.get('nama')
        kelas = request.form.get('kelas')  
        absen = request.form.get('absen')
        
        with sqlite3.connect(DATABASE) as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO absen (tgl, nama, kelas, absen) VALUES (?, ?, ?, ?)", (tgl, nama, kelas, absen))
            conn.commit()
        
        return redirect(url_for('absen_form'))  # Mengalihkan pengguna kembali ke rute 'dashboard'

@app.route('/edit_absen/<int:index>', methods=['GET', 'POST'])
def edit_absen(index):
    if request.method == 'POST':
        tgl = request.form.get('tgl')
        nama = request.form.get('nama')
        kelas = request.form.get('kelas')  
        absen = request.form.get('absen')

        with sqlite3.connect(DATABASE) as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE absen SET tgl=?, nama=?, kelas=?, absen=? WHERE id=?", (tgl, nama, kelas, absen, index))
            conn.commit()
        
        return redirect(url_for('dashboard'))
    return render_template('edit_absen.html', index=index)

@app.route('/delete_absen/<int:index>')
def delete_absen(index):
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM absen WHERE id=?", (index,))
        conn.commit()

    return redirect(url_for('dashboard'))

@app.route('/hafalan')
def hafalan_form():
    return render_template('input2.html')

@app.route('/add_hafalan', methods=['POST'])
def add_hafalan():
    if request.method == 'POST':
        tgl = request.form.get('tgl')
        nama = request.form.get('nama')  # Mengganti 'Nama' dengan 'nama'
        kelas = request.form.get('kelas')  # Mengganti 'Kelas' dengan 'kelas'
        surat = request.form.get('surat')
        ayat = request.form.get('ayat')
        ket = request.form.get('ket')
        
        with sqlite3.connect(DATABASE) as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO hafalan (tgl, nama, kelas, surat, ayat, ket) VALUES (?, ?, ?, ?, ?, ?)", (tgl, nama, kelas, surat, ayat, ket))
            conn.commit()

        return redirect(url_for('hafalan_form'))  # Mengalihkan ke dashboard setelah berhasil menyimpan data

    return "Metode yang digunakan bukan POST!"  # Menampilkan pesan jika metode yang digunakan bukan POST

@app.route('/edit_hafalan/<int:index>', methods=['GET', 'POST'])
def edit_hafalan(index):
    if request.method == 'POST':
        tgl = request.form.get('tgl')
        nama = request.form.get('nama')
        kelas = request.form.get('kelas')  
        surat = request.form.get('surat')
        ayat = request.form.get('ayat')  
        ket = request.form.get('ket')

        with sqlite3.connect(DATABASE) as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE hafalan SET tgl=?, nama=?, kelas=?, surat=?, ayat=?, ket=? WHERE id=?", (tgl, nama, kelas, surat, ayat, ket, index))
            conn.commit()
        return redirect(url_for('dashboard'))
    return render_template('edit_hafalan.html', index=index)

#delete hafalan
@app.route('/delete_hafalan/<int:id>')
def delete_hafalan(id):
    with sqlite3.connect(DATABASE) as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM hafalan WHERE id=?", (id,))
        conn.commit()
    return redirect(url_for('dashboard'))

@app.route('/export_to_excel')
def export_dashboard_to_excel():
    export_to_excel('data.db', 'data.xlsx', 'absen', 'hafalan')
    return send_file('data.xlsx', as_attachment=True)

def export_to_excel(database_file, excel_file, absen_table, hafalan_table):
    # Koneksi ke database SQLite
    conn = sqlite3.connect(database_file)
    cursor = conn.cursor()

    # Membuat workbook baru
    wb = Workbook()

    # Membuat sheet untuk tabel 'absen'
    ws_absen = wb.create_sheet(title='absen')

    # Mendapatkan data dari tabel 'absen'
    cursor.execute(f"SELECT * FROM {absen_table}")
    absen_rows = cursor.fetchall()

    # Menulis header kolom untuk sheet 'absen'
    columns = [description[0] for description in cursor.description]
    for col_index, col_name in enumerate(columns, start=1):
        ws_absen.cell(row=1, column=col_index, value=col_name)

    # Menulis data ke sheet 'absen'
    for row_index, row_data in enumerate(absen_rows, start=2):
        for col_index, col_data in enumerate(row_data, start=1):
            ws_absen.cell(row=row_index, column=col_index, value=col_data)

    # Membuat sheet untuk tabel 'hafalan'
    ws_hafalan = wb.create_sheet(title='hafalan')

    # Mendapatkan data dari tabel 'hafalan'
    cursor.execute(f"SELECT * FROM {hafalan_table}")
    hafalan_rows = cursor.fetchall()

    # Menulis header kolom untuk sheet 'hafalan'
    columns = [description[0] for description in cursor.description]
    for col_index, col_name in enumerate(columns, start=1):
        ws_hafalan.cell(row=1, column=col_index, value=col_name)

    # Menulis data ke sheet 'hafalan'
    for row_index, row_data in enumerate(hafalan_rows, start=2):
        for col_index, col_data in enumerate(row_data, start=1):
            ws_hafalan.cell(row=row_index, column=col_index, value=col_data)

    # Menyimpan workbook sebagai file Excel
    wb.save(excel_file)

    # Menutup koneksi ke database
    conn.close()


@app.route('/profile')
def profile():
    return render_template('profile.html')

if __name__ == '__main__':
    create_users_table()
    app.run(debug=True)
